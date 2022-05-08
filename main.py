import os
import openpyxl
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter


def excel_autowidth(path, target_excel:str, padding:int=0):
    """
    This is a minifunction that can be called independantly to auto-adjust width size of 
    multi-sheet Excel file. 

    :param: path: path to the directory that holds the Excel file.
    :param: target_excel: Excel file name, including `.xlsx`.
    :param: padding: not required, but adds padding to the column.
    :return: None
    """
    target_file = os.path.join(path, target_excel)
    wb = openpyxl.load_workbook(target_file)
    sheets = [sheet for sheet in wb.get_sheet_names()]

    for sheet in sheets:
        # Opening the worksheet:
        ws = wb[sheet]
        # Using dimension holder for original worksheet
        dim_holder = DimensionHolder(worksheet=ws)

        for col in range(ws.min_column, ws.max_column + 1):
            # intiating the width of the column as 0, as it will serve base size.
            width = 0
            for row in range(ws.min_row, ws.max_row + 1):
                cell_value = ws.cell(column=col, row=row).value
                if cell_value:
                    # finding what's the len of the cell's values and then setting it 
                    # as a new base `width`
                    cell_len = len(str(cell_value))
                    if cell_len > width:
                        width = cell_len + padding

            dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=width)

        ws.column_dimensions = dim_holder

    wb.save(target_file)
    print("Completed adjustments for {}".format(target_excel))


if __name__ == '__main__':
    pass
